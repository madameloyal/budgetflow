from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import os
import re
import io
import json
import unicodedata
from google.oauth2 import service_account
from googleapiclient.discovery import build

# ── Config ────────────────────────────────────────────────────────────────
SHEET_ID = os.environ.get("SHEET_ID", "1mc_ofkHODCLuhFlV4w8-D3c8KeMYPkIooYpQh0pqLAM")
SCOPES   = ["https://www.googleapis.com/auth/spreadsheets"]

# ── Multi-session config ──────────────────────────────────────────────────
# Each session = one event or budget scope, with its own prefixed tabs
SESSIONS = {
    "pa": {
        "name": "Paris XXL 2026",
        "prefix": "",  # legacy — no prefix (backward compatible)
        "dept_tabs": [
            "REPERAGES", "VHR", "PRODUCTION", "CATERING",
            "BAR", "ARTISTIQUE", "TECHNIQUE", "SCENO",
            "PARTENARIAT", "COMMUNICATION"
        ],
    },
    "bdx": {
        "name": "Bordeaux ML × Sonora 2026",
        "prefix": "BDX_",
        "dept_tabs": [
            "REPERAGES", "VHR", "PRODUCTION", "CATERING",
            "BAR", "ARTISTIQUE", "TECHNIQUE", "SCENO",
            "PARTENARIAT", "COMMUNICATION"
        ],
    },
    "epk": {
        "name": "Elektric Park 2026",
        "prefix": "EPK_",
        "dept_tabs": [
            "REPERAGES", "VHR", "PRODUCTION", "CATERING",
            "BAR", "ARTISTIQUE", "TECHNIQUE", "SCENO",
            "PARTENARIAT", "COMMUNICATION"
        ],
    },
    "gen": {
        "name": "Général — Frais de structure",
        "prefix": "GEN_",
        "dept_tabs": [
            "HUMAIN", "IMMOBILIER", "ASSURANCE", "JURIDIQUE",
            "DIVERS", "TRANSPORT", "RESTAURATION", "VHR",
            "COMMUNICATION", "URSAFF", "TAXES", "FOURNITURE_BUREAU",
            "ATELIER"
        ],
    },
}

DEFAULT_SESSION = "pa"

def get_session(session_id: str = None):
    """Return session config. Falls back to DEFAULT_SESSION."""
    sid = (session_id or DEFAULT_SESSION).lower().strip()
    if sid not in SESSIONS:
        raise HTTPException(status_code=400, detail=f"Session '{sid}' not found. Available: {list(SESSIONS.keys())}")
    return SESSIONS[sid]

def tab_name(session_cfg: dict, base_tab: str) -> str:
    """Apply session prefix to a tab name. e.g. 'BDX_' + 'PRODUCTION' = 'BDX_PRODUCTION'."""
    return f"{session_cfg['prefix']}{base_tab}"

# Legacy compat — keep DEPT_TABS for any code that still references it
DEPT_TABS = SESSIONS[DEFAULT_SESSION]["dept_tabs"]

RAW_TAB       = "QONTO_RAW"
UNMATCHED_TAB = "QONTO_UNMATCHED"
MATCH_LOG_TAB = "MATCH_LOG"
RECETTES_TAB  = "RECETTES"

MATCH_LOG_HEADERS = [
    "Dept", "Section", "Ligne", "Fournisseur", "Note", "Date",
    "HT", "TTC", "Source", "Timestamp"
]

COL = {
    "section": 0, "ligne": 1, "observations": 2, "typ": 3, "tva": 4,
    "est_ht": 5, "est_ttc": 6, "rh": 7, "reel_ht": 8, "reel_ttc": 9,
    "ecart": 10, "statut": 11
}

FIELD_TO_COL = {"est_ht": "F", "reel_ht": "I", "statut": "L", "observations": "C", "ligne": "B", "tva": "E", "typ": "D", "rh": "H"}

QC = {
    "cat": 0, "sous_cat": 1, "fourn": 2, "note": 3, "date": 4,
    "ht": 5, "ttc": 6, "ref": 7, "methode": 8, "init": 9, "carte": 10
}

POLE_MAP = {
    "PRODUCTION":   "PRODUCTION",
    "PROD":         "PRODUCTION",
    "TECHNIQUE":    "TECHNIQUE",
    "ARTISTIQUE":   "ARTISTIQUE",
    "VHR":          "VHR",
    "CATERING":     "CATERING",
    "RESTAURATION": "CATERING",
    "AMENAGEMENT":  "SCENO",
    "SCENO":        "SCENO",
    "COMMUNICATION":"COMMUNICATION",
    "PARTENARIAT":  "PARTENARIAT",
    "BAR":          "BAR",
    "REPERAGES":    "REPERAGES",
    "REPERAGE":     "REPERAGES",
    "SEMINAIRE":    "PRODUCTION",
}

SKIP_CATEGORIES  = {"HUMAIN", "PERFORMANCE", "DIVERS", "TRANSPORT", "LOGEMENT", "ATELIER"}
REVENUE_CATEGORY = "BILLETERIE"

RAW_HEADERS = [
    "Catégorie de trésorerie", "Sous-catégorie de trésorerie",
    "Nom de la contrepartie", "Note", "Date de l'opération (local)",
    "Montant total (HT)", "Montant total (TTC)", "Référence",
    "Méthode de paiement", "Initiateur", "Nom de la carte",
]

UNMATCHED_HEADERS = [
    "Catégorie", "Sous-catégorie", "Fournisseur", "Note", "Date",
    "Montant HT", "Montant TTC", "Référence",
    "Raison non-match", "Tab suggéré", "Ligne suggérée",
    "Statut", "Tab assigné", "Ligne assignée",
]

# ── Credentials ───────────────────────────────────────────────────────────
def get_creds_dict():
    raw = os.environ.get("GOOGLE_CREDENTIALS", "").strip().strip("'\"")
    if raw:
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            pass
        try:
            fixed = re.sub(
                r'("private_key"\s*:\s*")([\s\S]*?)(?="(?:\s*,|\s*}))',
                lambda m: m.group(1) + m.group(2).replace('\n', '\\n'),
                raw
            )
            return json.loads(fixed)
        except json.JSONDecodeError:
            pass
    private_key  = os.environ.get("PRIVATE_KEY", "").replace("\\n", "\n")
    client_email = os.environ.get("CLIENT_EMAIL", "")
    if private_key and client_email:
        return {
            "type": "service_account",
            "project_id": os.environ.get("PROJECT_ID", "budgetflow-489703"),
            "private_key_id": os.environ.get("PRIVATE_KEY_ID", ""),
            "private_key": private_key,
            "client_email": client_email,
            "client_id": os.environ.get("CLIENT_ID", ""),
            "token_uri": "https://oauth2.googleapis.com/token",
        }
    return None

def get_sheets_service():
    creds_dict = get_creds_dict()
    if not creds_dict:
        raise HTTPException(status_code=500, detail="No valid credentials found.")
    try:
        creds = service_account.Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        return build("sheets", "v4", credentials=creds).spreadsheets()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Credentials error: {str(e)}")

# ── Helpers ───────────────────────────────────────────────────────────────
def safe_float(v):
    try:
        s = str(v).replace(" ", "").replace("\u202f", "").replace("\u2009", "").replace(",", ".").replace("\u2212", "-")
        return float(s) if s not in ("", "—") else 0.0
    except:
        return 0.0

def normalize_str(s):
    if not s:
        return ""
    s = str(s).upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def fuzzy_match(a: str, b: str) -> bool:
    words_a = [w for w in a.split() if len(w) > 2]
    words_b = [w for w in b.split() if len(w) > 2]
    if not words_b:
        return False
    matches = sum(1 for w in words_b if any(w in wa or wa in w for wa in words_a))
    return (matches / len(words_b)) >= 0.6

def extract_note_desc(note: str):
    if not note:
        return None
    parts = note.split(" - ")
    if len(parts) >= 3:
        return " - ".join(parts[2:]).strip()
    slash = note.split(" / ")
    if len(slash) >= 3:
        return " / ".join(slash[2:]).strip()
    return None


def dept_from_note(note: str) -> str | None:
    """Extract department from note when Qonto category is empty.

    Handles three patterns found in real Qonto exports:
      1) "ARTISTIQUE - ML130326 - CRUSY AF + BF"  → ARTISTIQUE
      2) "Ml 130326 / aménagement / catering"      → AMENAGEMENT
      3) "ML.13&14MARS.PARIS.AMENAGEMENT.PEINTURE"  → AMENAGEMENT
      4) "Divers 2026 / partenariat Gallia / ..."   → PARTENARIAT
      5) "ML PARIS XXL ... - repérage - vhr - ..."  → VHR (or first recognized dept)
    """
    if not note:
        return None
    note_up = note.upper()
    n = normalize_str(note)

    # Pattern 1: "DEPT - MLxxxxxx - DESCRIPTION"
    parts = note.split(" - ")
    if len(parts) >= 2 and "ML" in parts[1].upper():
        candidate = normalize_str(parts[0])
        if candidate and POLE_MAP.get(candidate):
            return POLE_MAP[candidate]

    # Pattern 2: "Ml 130326 / dept / detail"
    slash_parts = note.split("/")
    if len(slash_parts) >= 2:
        for sp in slash_parts[1:]:
            candidate = normalize_str(sp.strip())
            if candidate and POLE_MAP.get(candidate):
                return POLE_MAP[candidate]

    # Pattern 3: "ML.13&14MARS.PARIS.AMENAGEMENT.PEINTURE" — dot-separated
    if "ML" in note_up and "." in note:
        dot_parts = note_up.replace("&", ".").split(".")
        for dp in dot_parts:
            candidate = normalize_str(dp.strip())
            if candidate and POLE_MAP.get(candidate):
                return POLE_MAP[candidate]

    # Pattern 4: "Divers 2026 / partenariat Gallia / ..." — keyword in any slash segment
    # Pattern 5: Generic fallback — scan all words for a known dept
    for word in n.split():
        if word and len(word) > 3 and POLE_MAP.get(word):
            return POLE_MAP[word]

    return None

def is_total_row(section: str, ligne: str) -> bool:
    return "TOTAL" in (section + " " + ligne).upper()

def parse_dept_rows(values: list, tab: str) -> list:
    rows = []
    current_section = ""
    for i, row in enumerate(values):
        if i < 2:
            continue
        row = list(row) + [""] * (12 - len(row))
        section_val = str(row[COL["section"]]).strip()
        ligne_val   = str(row[COL["ligne"]]).strip()
        if section_val and not ligne_val:
            current_section = section_val
            continue
        if "TOTAL" in ligne_val.upper():
            continue
        if not ligne_val and not any(str(c).strip() for c in row[5:10]):
            continue
        rows.append({
            "section":      current_section,
            "ligne":        ligne_val or f"(ligne {i})",
            "observations": str(row[COL["observations"]]).strip(),
            "typ":          str(row[COL["typ"]]).strip(),
            "tva":          str(row[COL["tva"]]).strip(),
            "est_ht":       safe_float(row[COL["est_ht"]]),
            "rh":           str(row[COL["rh"]]).strip(),
            "reel_ht":      safe_float(row[COL["reel_ht"]]),
            "statut":       str(row[COL["statut"]]).strip(),
            "_row":         i + 1
        })
    return rows

# ── App ───────────────────────────────────────────────────────────────────
app = FastAPI(title="BudgetFlow API", version="2.1")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

@app.get("/")
def health():
    creds = get_creds_dict()
    return {
        "status": "ok",
        "service": "BudgetFlow API",
        "version": "2.1",
        "endpoints": ["/api/sessions", "/api/budget", "/api/unmatched", "/api/assign", "/api/import-qonto"],
        "credentials_loaded": creds is not None,
        "credential_method": "json" if os.environ.get("GOOGLE_CREDENTIALS") else "individual_vars"
    }

@app.get("/api/sessions")
def list_sessions():
    """Return available budget sessions."""
    return {
        "status": "ok",
        "sessions": {
            sid: {"name": s["name"], "prefix": s["prefix"], "dept_count": len(s["dept_tabs"]), "dept_tabs": s["dept_tabs"]}
            for sid, s in SESSIONS.items()
        }
    }

@app.get("/api/budget")
def get_budget(session: str = None):
    sess = get_session(session)
    try:
        svc = get_sheets_service()
        result = {}
        for base_tab in sess["dept_tabs"]:
            full_tab = tab_name(sess, base_tab)
            resp = svc.values().get(
                spreadsheetId=SHEET_ID,
                range=f"{full_tab}!A:L",
                valueRenderOption="UNFORMATTED_VALUE"
            ).execute()
            result[base_tab] = parse_dept_rows(resp.get("values", []), base_tab)
        return {"status": "ok", "session": session or DEFAULT_SESSION, "data": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/budget/{dept}")
def get_dept(dept: str, session: str = None):
    sess = get_session(session)
    base_tab = dept.upper()
    if base_tab not in list(sess["dept_tabs"]) + [RECETTES_TAB]:
        raise HTTPException(status_code=404, detail=f"Dept '{dept}' not found in session '{session or DEFAULT_SESSION}'")
    try:
        svc = get_sheets_service()
        full_tab = tab_name(sess, base_tab)
        resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{full_tab}!A:L",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        return {"status": "ok", "dept": base_tab, "data": parse_dept_rows(resp.get("values", []), base_tab)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class UpdateCell(BaseModel):
    dept:  str
    row:   int
    field: str
    value: str
    session: str = None

@app.post("/api/budget/update")
def update_cell(payload: UpdateCell):
    sess = get_session(payload.session)
    base_tab = payload.dept.upper()
    valid_tabs = list(sess["dept_tabs"]) + [RECETTES_TAB]
    if base_tab not in valid_tabs:
        raise HTTPException(status_code=404, detail=f"Dept '{base_tab}' not found in session")
    full_tab = tab_name(sess, base_tab)
    col = FIELD_TO_COL.get(payload.field)
    if not col:
        raise HTTPException(status_code=400, detail=f"Field '{payload.field}' not writable")
    try:
        svc = get_sheets_service()
        cell_range = f"{full_tab}!{col}{payload.row}"
        val = payload.value
        if payload.field in ("est_ht", "reel_ht"):
            try:
                val = float(val)
            except:
                pass
        svc.values().update(
            spreadsheetId=SHEET_ID,
            range=cell_range,
            valueInputOption="USER_ENTERED",
            body={"values": [[val]]}
        ).execute()
        return {"status": "ok", "updated": cell_range, "value": val}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class AddLigne(BaseModel):
    dept:    str
    section: str
    ligne:   str
    obs:     str = ""
    typ:     str = ""
    tva:     str = ""
    est_ht:  float = 0.0
    statut:  str = "ESTIMATION"
    session: str = None

@app.post("/api/budget/add-ligne")
def add_ligne(payload: AddLigne):
    sess = get_session(payload.session)
    base_tab = payload.dept.upper()
    if base_tab not in list(sess["dept_tabs"]) + [RECETTES_TAB]:
        raise HTTPException(status_code=404, detail=f"Dept \'{base_tab}\' not found in session")
    tab = tab_name(sess, base_tab)
    try:
        svc = get_sheets_service()
        # Read current sheet to find where to insert (after last row of the section)
        resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{tab}!A:L",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        values = resp.get("values", [])

        # Find the last row of the target section
        insert_after = len(values)  # default: append at end
        in_section = False
        for i, row in enumerate(values):
            row_p = list(row) + [""] * 12
            sec_val   = str(row_p[COL["section"]]).strip().upper()
            ligne_val = str(row_p[COL["ligne"]]).strip().upper()
            target_sec = payload.section.strip().upper()

            if sec_val == target_sec and not ligne_val:
                in_section = True
            if in_section:
                # Stop at TOTAL row or next section header
                if "TOTAL" in ligne_val or ("TOTAL" in sec_val and not ligne_val and sec_val != target_sec):
                    insert_after = i  # insert before this row
                    break
                # Check if we've moved to a new section
                if sec_val and sec_val != target_sec and not ligne_val and i > 2:
                    insert_after = i
                    break
                insert_after = i + 1

        # Build the new row (12 cols: A-L)
        new_row = [""] * 12
        new_row[COL["section"]]      = ""  # section is a header row, ligne rows leave it blank
        new_row[COL["ligne"]]        = payload.ligne
        new_row[COL["observations"]] = payload.obs
        new_row[COL["typ"]]          = payload.typ
        new_row[COL["tva"]]          = payload.tva
        new_row[COL["est_ht"]]       = payload.est_ht if payload.est_ht else ""
        new_row[COL["reel_ht"]]      = 0
        new_row[COL["reel_ttc"]]     = 0
        new_row[COL["statut"]]       = payload.statut

        # Get sheet ID for insert
        sheet_meta = svc.get(spreadsheetId=SHEET_ID).execute()
        sheet_id = next(
            (s["properties"]["sheetId"] for s in sheet_meta["sheets"]
             if s["properties"]["title"] == tab), None
        )
        if sheet_id is None:
            raise HTTPException(status_code=404, detail=f"Sheet '{tab}' not found")

        # Insert a blank row at insert_after position
        svc.batchUpdate(
            spreadsheetId=SHEET_ID,
            body={"requests": [{"insertDimension": {
                "range": {"sheetId": sheet_id, "dimension": "ROWS",
                          "startIndex": insert_after, "endIndex": insert_after + 1},
                "inheritFromBefore": True
            }}]}
        ).execute()

        # Write the new row data
        sheet_row = insert_after + 1  # 1-based
        svc.values().update(
            spreadsheetId=SHEET_ID,
            range=f"{tab}!A{sheet_row}:L{sheet_row}",
            valueInputOption="USER_ENTERED",
            body={"values": [new_row]}
        ).execute()

        return {"status": "ok", "inserted_row": sheet_row, "ligne": payload.ligne, "dept": tab}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── QONTO_UNMATCHED ────────────────────────────────────────────────────────
UNMATCHED_COLS = {"cat": 0, "sous_cat": 1, "fourn": 2, "note": 3, "date": 4, "ht": 5, "ttc": 6, "ref": 7, "raison": 8, "tab": 9, "ligne": 10, "statut": 11, "init": 12}

def parse_unmatched_rows(values: list) -> list:
    rows = []
    for i, row in enumerate(values):
        if i == 0:
            continue
        row = list(row) + [""] * (14 - len(row))
        if not any(str(c).strip() for c in row):
            continue
        rows.append({
            "idx":   i,
            "dept":  str(row[UNMATCHED_COLS["cat"]]).strip(),
            "section": str(row[UNMATCHED_COLS["sous_cat"]]).strip(),
            "fourn": str(row[UNMATCHED_COLS["fourn"]]).strip(),
            "note":  str(row[UNMATCHED_COLS["note"]]).strip(),
            "date":  str(row[UNMATCHED_COLS["date"]]).strip(),
            "ht":    safe_float(row[UNMATCHED_COLS["ht"]]),
            "ttc":   safe_float(row[UNMATCHED_COLS["ttc"]]),
            "ref":   str(row[UNMATCHED_COLS["ref"]]).strip(),
            "raison": str(row[UNMATCHED_COLS["raison"]]).strip(),
            "init":  str(row[UNMATCHED_COLS.get("init", 12)]).strip(),
        })
    return rows

@app.get("/api/unmatched")
def get_unmatched(session: str = None):
    sess = get_session(session)
    unmatched_tab = tab_name(sess, UNMATCHED_TAB)
    try:
        svc = get_sheets_service()
        resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{unmatched_tab}!A:N",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        rows = parse_unmatched_rows(resp.get("values", []))
        return {"status": "ok", "count": len(rows), "data": rows}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/qonto-raw")
def get_qonto_raw(session: str = None):
    """Return all rows from QONTO_RAW tab for the Relevé Qonto dashboard tab."""
    sess = get_session(session)
    raw_tab = tab_name(sess, RAW_TAB)
    try:
        svc = get_sheets_service()
        resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{raw_tab}!A:K",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        values = resp.get("values", [])
        if len(values) < 2:
            return {"status": "ok", "count": 0, "data": []}
        rows = []
        for i, row in enumerate(values):
            if i == 0:
                continue  # skip header
            row = list(row) + [""] * (11 - len(row))
            if not any(str(c).strip() for c in row):
                continue
            rows.append({
                "cat":   str(row[QC["cat"]]).strip(),
                "fourn": str(row[QC["fourn"]]).strip(),
                "note":  str(row[QC["note"]]).strip(),
                "date":  str(row[QC["date"]]).strip(),
                "ht":    safe_float(row[QC["ht"]]),
                "ttc":   safe_float(row[QC["ttc"]]),
                "ref":   str(row[QC["ref"]]).strip(),
                "init":  str(row[QC["init"]]).strip(),
            })
        return {"status": "ok", "count": len(rows), "data": rows}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── RECETTES ─────────────────────────────────────────────────────────────

@app.get("/api/recettes")
def get_recettes(session: str = None):
    """Return recettes data from the RECETTES tab.

    Uses same column structure as dept tabs (A:L):
    Section=cat (TICKETS/BAR/DIVERS/AUTRE), Ligne, Observations, Type, TVA,
    Est HT, Est TTC, RH (=QTE), Réel HT, Réel TTC, Écart, Statut.

    Note: column H (RH) is repurposed as QTE for recettes, and
    Observations (C) stores Prix Unitaire as a number.
    """
    sess = get_session(session)
    rec_tab = tab_name(sess, RECETTES_TAB)
    try:
        svc = get_sheets_service()
        resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{rec_tab}!A:L",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        values = resp.get("values", [])
        if len(values) < 2:
            return {"status": "ok", "data": []}

        rows = parse_dept_rows(values, "RECETTES")
        # Enrich with qte and pu from the repurposed columns
        for row in rows:
            row["qte"] = safe_float(row.get("rh", 0))
            row["pu"] = safe_float(row.get("observations", 0))
            # Clean observations — if it was used for PU, don't show as text
            try:
                float(str(row.get("observations", "")).replace(",", "."))
                row["observations"] = ""
            except (ValueError, TypeError):
                pass
        return {"status": "ok", "data": rows}
    except HTTPException:
        raise
    except Exception as e:
        # Tab might not exist yet — return empty
        return {"status": "ok", "data": []}

# ── MATCH LOG ─────────────────────────────────────────────────────────────

def ensure_match_log_tab(svc, ml_tab=None):
    """Create MATCH_LOG tab if it doesn't exist, with headers."""
    if ml_tab is None:
        ml_tab = MATCH_LOG_TAB
    try:
        meta = svc.get(spreadsheetId=SHEET_ID).execute()
        existing = {s["properties"]["title"] for s in meta["sheets"]}
        if ml_tab not in existing:
            svc.batchUpdate(
                spreadsheetId=SHEET_ID,
                body={"requests": [{"addSheet": {"properties": {"title": ml_tab}}}]}
            ).execute()
            svc.values().update(
                spreadsheetId=SHEET_ID, range=f"{ml_tab}!A1",
                valueInputOption="RAW",
                body={"values": [MATCH_LOG_HEADERS]}
            ).execute()
    except Exception as e:
        print(f"Warning: could not ensure MATCH_LOG tab: {e}")

def append_match_log(svc, dept: str, section: str, ligne: str, fourn: str, note: str, date: str, ht: float, ttc: float, source: str, ml_tab: str = None):
    """Append a row to the MATCH_LOG tab."""
    from datetime import datetime
    if ml_tab is None:
        ml_tab = MATCH_LOG_TAB
    try:
        ensure_match_log_tab(svc, ml_tab)
        timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        svc.values().append(
            spreadsheetId=SHEET_ID,
            range=f"{ml_tab}!A:J",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body={"values": [[dept, section, ligne, fourn, note, date, ht, ttc, source, timestamp]]}
        ).execute()
    except Exception as e:
        print(f"Warning: could not append to MATCH_LOG: {e}")

def build_tx_fingerprint(date: str, fourn: str, ttc, ref: str) -> str:
    """Build a dedup key for a Qonto transaction: date|fourn|ttc|ref."""
    return f"{str(date).strip()}|{normalize_str(fourn)}|{safe_float(ttc)}|{str(ref).strip()}"


def recalc_reel_from_match_log(svc, sess):
    """Recalculate all RÉEL columns from MATCH_LOG (the single source of truth).

    Reads the full MATCH_LOG, sums HT/TTC per dept+ligne, finds the
    corresponding row in each dept sheet, and writes the totals.
    """
    ml_tab = tab_name(sess, MATCH_LOG_TAB)
    s_prefix = sess["prefix"]

    # 1. Read full MATCH_LOG
    try:
        log_resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{ml_tab}!A:J",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        log_rows = log_resp.get("values", [])
    except Exception:
        log_rows = []

    # 2. Sum per (dept_tab, ligne) — skip header row
    sums = {}  # { (dept_tab, ligne): [ht, ttc] }
    for i, row in enumerate(log_rows):
        if i == 0:
            continue
        row = list(row) + [""] * 10
        dept_tab = str(row[0]).strip()
        ligne = str(row[2]).strip()
        ht = abs(safe_float(row[6]))
        ttc = abs(safe_float(row[7]))
        if not dept_tab or not ligne:
            continue
        key = (dept_tab, ligne)
        if key not in sums:
            sums[key] = [0.0, 0.0]
        sums[key][0] += ht
        sums[key][1] += ttc

    # 3. For each dept tab, read rows and build a write batch
    write_data = []
    for base_t in sess["dept_tabs"]:
        full_t = f"{s_prefix}{base_t}"
        try:
            resp = svc.values().get(
                spreadsheetId=SHEET_ID,
                range=f"{full_t}!A:L",
                valueRenderOption="UNFORMATTED_VALUE"
            ).execute()
            vals = resp.get("values", [])
        except Exception:
            continue

        for i, row in enumerate(vals):
            if i < 2:
                continue
            rp = list(row) + [""] * 12
            ligne = str(rp[COL["ligne"]]).strip()
            section = str(rp[COL["section"]]).strip()
            if not ligne or ligne.startswith("(ligne") or is_total_row(section, ligne):
                continue
            # Look up in sums — try both with and without prefix
            key_full = (full_t, ligne)
            key_base = (base_t, ligne)
            total = sums.get(key_full) or sums.get(key_base) or [0.0, 0.0]
            write_data.append({
                "range": f"{full_t}!I{i+1}:J{i+1}",
                "values": [[total[0], total[1]]]
            })

    # 4. Write all RÉEL values in one batch
    if write_data:
        svc.values().batchUpdate(
            spreadsheetId=SHEET_ID,
            body={"valueInputOption": "RAW", "data": write_data}
        ).execute()


@app.get("/api/match-log")
def get_match_log(dept: str = "", ligne: str = "", session: str = None):
    """Return match log rows, optionally filtered by dept and/or ligne."""
    sess = get_session(session)
    ml_tab = tab_name(sess, MATCH_LOG_TAB)
    try:
        svc = get_sheets_service()
        ensure_match_log_tab(svc, ml_tab)
        resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{ml_tab}!A:J",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        values = resp.get("values", [])
        rows = []
        for i, row in enumerate(values):
            if i == 0:
                continue
            row = list(row) + [""] * (10 - len(row))
            entry = {
                "dept": str(row[0]).strip(),
                "section": str(row[1]).strip(),
                "ligne": str(row[2]).strip(),
                "fourn": str(row[3]).strip(),
                "note": str(row[4]).strip(),
                "date": str(row[5]).strip(),
                "ht": safe_float(row[6]),
                "ttc": safe_float(row[7]),
                "source": str(row[8]).strip(),
                "timestamp": str(row[9]).strip(),
            }
            if dept and entry["dept"].upper() != dept.upper():
                continue
            if ligne and entry["ligne"].strip() != ligne.strip():
                continue
            rows.append(entry)
        return {"status": "ok", "count": len(rows), "data": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── MANUAL ASSIGN ──────────────────────────────────────────────────────────
class AssignPayload(BaseModel):
    unmatched_idx: int
    dept:          str
    ligne:         str
    session:       str = None

@app.post("/api/assign")
def assign_transaction(payload: AssignPayload):
    sess = get_session(payload.session)
    base_tab = payload.dept.upper()
    if base_tab not in sess["dept_tabs"]:
        raise HTTPException(status_code=404, detail=f"Dept '{base_tab}' not found in session")
    tab = tab_name(sess, base_tab)
    unmatched_tab = tab_name(sess, UNMATCHED_TAB)
    ml_tab = tab_name(sess, MATCH_LOG_TAB)
    try:
        svc = get_sheets_service()
        unmatched_resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{unmatched_tab}!A:N",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        unmatched_rows = unmatched_resp.get("values", [])
        if payload.unmatched_idx >= len(unmatched_rows):
            raise HTTPException(status_code=404, detail="Unmatched row not found")
        tx_row = list(unmatched_rows[payload.unmatched_idx]) + [""] * 14
        ht_val  = safe_float(tx_row[UNMATCHED_COLS["ht"]])
        ttc_val = safe_float(tx_row[UNMATCHED_COLS["ttc"]])

        # Verify budget line exists
        dept_resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{tab}!A:L",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        dept_rows = dept_resp.get("values", [])
        line_exists = False
        for i, row in enumerate(dept_rows):
            row = list(row) + [""] * 12
            if str(row[COL["ligne"]]).strip() == payload.ligne.strip():
                line_exists = True
                break
        if not line_exists:
            raise HTTPException(status_code=404, detail=f"Ligne '{payload.ligne}' not found in {tab}")

        # 1. Remove from QONTO_UNMATCHED
        sheet_meta = svc.get(spreadsheetId=SHEET_ID).execute()
        unmatched_sheet_id = next(
            (s["properties"]["sheetId"] for s in sheet_meta["sheets"]
             if s["properties"]["title"] == unmatched_tab), None
        )
        if unmatched_sheet_id is not None:
            svc.batchUpdate(
                spreadsheetId=SHEET_ID,
                body={"requests": [{"deleteDimension": {"range": {
                    "sheetId": unmatched_sheet_id, "dimension": "ROWS",
                    "startIndex": payload.unmatched_idx,
                    "endIndex": payload.unmatched_idx + 1
                }}}]}
            ).execute()

        # 2. Append to MATCH_LOG
        tx_fourn = str(tx_row[UNMATCHED_COLS["fourn"]]).strip()
        tx_note  = str(tx_row[UNMATCHED_COLS["note"]]).strip()
        tx_date  = str(tx_row[UNMATCHED_COLS["date"]]).strip()
        tx_section = str(tx_row[UNMATCHED_COLS["sous_cat"]]).strip()
        append_match_log(svc, tab, tx_section, payload.ligne, tx_fourn, tx_note, tx_date, abs(ht_val), abs(ttc_val), "manual", ml_tab)

        # 3. Recalculate ALL réel from MATCH_LOG
        recalc_reel_from_match_log(svc, sess)

        return {"status": "ok", "assigned": payload.ligne, "dept": tab,
                "ht_added": abs(ht_val), "ttc_added": abs(ttc_val)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── DISSOCIATE (undo a match) ─────────────────────────────────────────────
class DissociatePayload(BaseModel):
    session: str = None
    dept:   str
    ligne:  str
    fourn:  str
    note:   str
    ht:     float
    ttc:    float

@app.post("/api/dissociate")
def dissociate_transaction(payload: DissociatePayload):
    sess = get_session(payload.session)
    base_tab = payload.dept.upper()
    if base_tab not in sess["dept_tabs"]:
        raise HTTPException(status_code=404, detail=f"Dept '{base_tab}' not found in session")
    tab = tab_name(sess, base_tab)
    unmatched_tab = tab_name(sess, UNMATCHED_TAB)
    ml_tab = tab_name(sess, MATCH_LOG_TAB)
    try:
        svc = get_sheets_service()

        # 1. Remove matching row from MATCH_LOG
        ensure_match_log_tab(svc, ml_tab)
        log_resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{ml_tab}!A:J",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        log_rows = log_resp.get("values", [])
        log_row_to_delete = None
        for i, row in enumerate(log_rows):
            if i == 0:
                continue
            row = list(row) + [""] * 10
            if (str(row[0]).strip().upper() == tab.upper() and
                str(row[2]).strip() == payload.ligne.strip() and
                str(row[3]).strip() == payload.fourn.strip() and
                abs(safe_float(row[7]) - abs(payload.ttc)) < 0.01):
                log_row_to_delete = i
                break

        if log_row_to_delete is not None:
            sheet_meta = svc.get(spreadsheetId=SHEET_ID).execute()
            log_sheet_id = next(
                (s["properties"]["sheetId"] for s in sheet_meta["sheets"]
                 if s["properties"]["title"] == ml_tab), None
            )
            if log_sheet_id is not None:
                svc.batchUpdate(
                    spreadsheetId=SHEET_ID,
                    body={"requests": [{"deleteDimension": {"range": {
                        "sheetId": log_sheet_id, "dimension": "ROWS",
                        "startIndex": log_row_to_delete,
                        "endIndex": log_row_to_delete + 1
                    }}}]}
                ).execute()

        # 2. Re-add to QONTO_UNMATCHED
        unmatched_row = [
            base_tab, "", payload.fourn, payload.note, "",
            abs(payload.ht), abs(payload.ttc), "",
            "Dissocié manuellement", "", "", "PENDING", "", ""
        ]
        svc.values().append(
            spreadsheetId=SHEET_ID,
            range=f"{unmatched_tab}!A:N",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body={"values": [unmatched_row]}
        ).execute()

        # 3. Recalculate ALL réel from MATCH_LOG
        recalc_reel_from_match_log(svc, sess)

        return {"status": "ok", "dissociated": payload.ligne, "dept": tab,
                "ht_removed": abs(payload.ht), "ttc_removed": abs(payload.ttc)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ══════════════════════════════════════════════════════════════════════════
# QONTO IMPORT — file upload directly to Railway (bypasses Apps Script 50KB limit)
# ══════════════════════════════════════════════════════════════════════════

def parse_qonto_file(file_bytes: bytes, filename: str) -> list:
    import csv
    fname = filename.lower()

    if fname.endswith(".csv") or fname.endswith(".tsv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        sep  = "\t" if "\t" in text.split("\n")[0] else ","
        reader = csv.reader(io.StringIO(text), delimiter=sep)
        all_rows = list(reader)
        return [list(map(str, r)) for r in all_rows[1:] if any(c.strip() for c in r)]

    if fname.endswith(".xls"):
        try:
            import xlrd
            # Try utf-8 first, fall back to latin-1 if needed
            for enc in ('utf-8', 'latin-1', 'cp1252'):
                try:
                    wb = xlrd.open_workbook(file_contents=file_bytes, encoding_override=enc)
                    ws = wb.sheet_by_index(0)
                    # Quick check: if first row has mojibake, try next encoding
                    sample = ' '.join(str(ws.cell(0,j).value) for j in range(min(ws.ncols,5)))
                    if 'Ã' in sample or '\x83' in sample:
                        continue
                    break
                except Exception:
                    continue
            ws = wb.sheet_by_index(0)
            rows = []
            for i in range(1, ws.nrows):
                row = []
                for j in range(ws.ncols):
                    cell = ws.cell(i, j)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        import datetime
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row.append(dt.strftime("%Y/%m/%d"))
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        v = cell.value
                        row.append(str(int(v)) if v == int(v) else str(v))
                    else:
                        val = str(cell.value).strip()
                        # Fix mojibake if still present: re-encode as latin-1, decode as utf-8
                        try:
                            fixed = val.encode('latin-1').decode('utf-8')
                            val = fixed
                        except (UnicodeEncodeError, UnicodeDecodeError):
                            pass
                        row.append(val)
                if any(c.strip() for c in row):
                    rows.append(row)
            return rows
        except ImportError:
            pass  # fall through to openpyxl

    # xlsx or xls fallback
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            str_row = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in str_row):
                rows.append(str_row)
        return rows
    except ImportError:
        raise HTTPException(status_code=500, detail="No XLS parser available on server. Install xlrd or openpyxl.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")


def find_budget_line(dept_vals: list, note: str, fourn: str, ref: str):
    candidates = []
    for i, row in enumerate(dept_vals):
        row = list(row) + [""] * 12
        ligne   = str(row[COL["ligne"]]).strip()
        section = str(row[COL["section"]]).strip()
        obs     = str(row[COL["observations"]]).strip()
        if not ligne or ligne.startswith("(ligne") or is_total_row(section, ligne):
            continue
        candidates.append({"row": i + 1, "ligne": ligne, "section": section, "obs": obs})

    if not candidates:
        return None

    note_desc = extract_note_desc(note)
    if note_desc:
        nd = normalize_str(note_desc)
        for c in candidates:
            if normalize_str(c["ligne"]) == nd:
                return c
        for c in candidates:
            if fuzzy_match(normalize_str(c["ligne"]), nd):
                return c

    if fourn:
        nf = normalize_str(fourn)
        for c in candidates:
            nl = normalize_str(c["ligne"])
            if nl and nf and (nl in nf or nf in nl):
                return c

    if ref:
        for c in candidates:
            if c["obs"] and ref in c["obs"]:
                return c

    return None


@app.post("/api/import-qonto")
async def import_qonto(file: UploadFile = File(...), session: str = None):
    try:
        file_bytes = await file.read()
        filename   = file.filename or "upload.xls"
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File read error: {e}")

    sess = get_session(session)
    raw_tab = tab_name(sess, RAW_TAB)
    unmatched_tab = tab_name(sess, UNMATCHED_TAB)
    ml_tab = tab_name(sess, MATCH_LOG_TAB)
    s_dept_tabs = sess["dept_tabs"]
    s_prefix = sess["prefix"]

    try:
        raw_rows = parse_qonto_file(file_bytes, filename)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Parse error: {e}")

    if not raw_rows:
        raise HTTPException(status_code=400, detail="Aucune transaction trouvée dans le fichier.")

    svc = get_sheets_service()

    # Ensure tabs exist
    try:
        sheet_meta = svc.get(spreadsheetId=SHEET_ID).execute()
        existing_tabs = {s["properties"]["title"]: s["properties"]["sheetId"]
                         for s in sheet_meta["sheets"]}

        tabs_to_create = [t for t in [raw_tab, unmatched_tab, ml_tab] if t not in existing_tabs]
        if tabs_to_create:
            svc.batchUpdate(
                spreadsheetId=SHEET_ID,
                body={"requests": [{"addSheet": {"properties": {"title": t}}} for t in tabs_to_create]}
            ).execute()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Tab setup error: {e}")

    # ── STEP 1: Read existing QONTO_RAW to build dedup set ──
    existing_fingerprints = set()
    try:
        existing_raw_resp = svc.values().get(
            spreadsheetId=SHEET_ID,
            range=f"{raw_tab}!A:K",
            valueRenderOption="UNFORMATTED_VALUE"
        ).execute()
        existing_raw_rows = existing_raw_resp.get("values", [])
        for i, row in enumerate(existing_raw_rows):
            if i == 0:
                continue  # skip header
            row = list(row) + [""] * 11
            fp = build_tx_fingerprint(row[QC["date"]], row[QC["fourn"]], row[QC["ttc"]], row[QC["ref"]])
            existing_fingerprints.add(fp)
    except Exception:
        existing_raw_rows = []

    # ── STEP 2: Filter incoming rows — keep only NEW transactions ──
    new_raw_rows = []
    duplicate_count = 0
    for raw in raw_rows:
        raw = list(raw) + [""] * max(0, 11 - len(raw))
        fp = build_tx_fingerprint(raw[QC["date"]], raw[QC["fourn"]], raw[QC["ttc"]], raw[QC["ref"]])
        if fp in existing_fingerprints:
            duplicate_count += 1
        else:
            new_raw_rows.append(raw)
            existing_fingerprints.add(fp)  # prevent dupes within same file

    # ── STEP 3: Append new rows to QONTO_RAW (don't clear) ──
    if new_raw_rows:
        try:
            raw_append = [
                (list(r) + [""] * len(RAW_HEADERS))[:len(RAW_HEADERS)] for r in new_raw_rows
            ]
            # If sheet was empty (no existing rows), write headers first
            if not existing_raw_rows:
                svc.values().update(
                    spreadsheetId=SHEET_ID, range=f"{raw_tab}!A1",
                    valueInputOption="RAW", body={"values": [RAW_HEADERS]}
                ).execute()
            svc.values().append(
                spreadsheetId=SHEET_ID,
                range=f"{raw_tab}!A:K",
                valueInputOption="RAW",
                insertDataOption="INSERT_ROWS",
                body={"values": raw_append}
            ).execute()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error appending QONTO_RAW: {e}")

    # ── STEP 4: Load dept data for matching (NO réel reset) ──
    dept_data_cache = {}
    try:
        for base_t in s_dept_tabs:
            full_t = f"{s_prefix}{base_t}"
            resp = svc.values().get(
                spreadsheetId=SHEET_ID, range=f"{full_t}!A:L",
                valueRenderOption="UNFORMATTED_VALUE"
            ).execute()
            dept_data_cache[base_t] = resp.get("values", [])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading dept data: {e}")

    # ── STEP 5: Auto-match only NEW transactions ──
    results = {"matched": 0, "unmatched": 0, "revenue": 0, "skipped": 0, "duplicates": duplicate_count}
    new_unmatched_rows = []
    new_match_log_rows = []

    norm_skip = {normalize_str(s) for s in SKIP_CATEGORIES}
    norm_rev  = normalize_str(REVENUE_CATEGORY)

    for raw in new_raw_rows:
        raw = list(raw) + [""] * max(0, 11 - len(raw))
        cat   = normalize_str(raw[QC["cat"]])
        fourn = raw[QC["fourn"]].strip()
        note  = raw[QC["note"]].strip()
        date  = raw[QC["date"]].strip()
        ref   = raw[QC["ref"]].strip()
        ttc   = safe_float(raw[QC["ttc"]])
        ht    = safe_float(raw[QC["ht"]])
        base  = (list(raw) + [""] * len(RAW_HEADERS))[:len(RAW_HEADERS)]

        if cat == norm_rev:
            results["revenue"] += 1
            continue

        if cat in norm_skip:
            new_unmatched_rows.append(base + [f'Catégorie "{cat}" non mappée', "", "", "PENDING", "", ""])
            results["unmatched"] += 1
            continue

        if ttc > 0:
            results["skipped"] += 1
            continue

        target_tab = POLE_MAP.get(cat)

        # ── Note-based fallback: when Qonto category is empty, parse the note ──
        if not target_tab and not cat:
            note_dept = dept_from_note(note)
            if note_dept:
                target_tab = note_dept

        if not target_tab:
            reason = f'Catégorie "{cat}" inconnue' if cat else "Pas de catégorie ni de département dans la note"
            new_unmatched_rows.append(base + [reason, "", "", "PENDING", "", ""])
            results["unmatched"] += 1
            continue

        match = find_budget_line(dept_data_cache.get(target_tab, []), note, fourn, ref)

        if match:
            new_match_log_rows.append([target_tab, match.get("section",""), match["ligne"], fourn, note, date, abs(ht), abs(ttc), "auto"])
            results["matched"] += 1
        else:
            new_unmatched_rows.append(base + ["Ligne budgétaire non trouvée", target_tab, "", "PENDING", "", ""])
            results["unmatched"] += 1

    # ── STEP 6: Append new matches to MATCH_LOG (don't clear) ──
    if new_match_log_rows:
        try:
            from datetime import datetime
            ensure_match_log_tab(svc, ml_tab)
            timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            log_append = [row + [timestamp] for row in new_match_log_rows]
            svc.values().append(
                spreadsheetId=SHEET_ID,
                range=f"{ml_tab}!A:J",
                valueInputOption="RAW",
                insertDataOption="INSERT_ROWS",
                body={"values": log_append}
            ).execute()
        except Exception as e:
            print(f"Warning: could not append to MATCH_LOG: {e}")

    # ── STEP 7: Append new unmatched rows (don't clear) ──
    if new_unmatched_rows:
        try:
            unmatched_append = [
                (list(r) + [""] * len(UNMATCHED_HEADERS))[:len(UNMATCHED_HEADERS)]
                for r in new_unmatched_rows
            ]
            # If sheet was empty, write headers first
            try:
                existing_unmatched = svc.values().get(
                    spreadsheetId=SHEET_ID, range=f"{unmatched_tab}!A1",
                    valueRenderOption="UNFORMATTED_VALUE"
                ).execute()
                has_unmatched_data = bool(existing_unmatched.get("values"))
            except Exception:
                has_unmatched_data = False
            if not has_unmatched_data:
                svc.values().update(
                    spreadsheetId=SHEET_ID, range=f"{unmatched_tab}!A1",
                    valueInputOption="RAW", body={"values": [UNMATCHED_HEADERS]}
                ).execute()
            svc.values().append(
                spreadsheetId=SHEET_ID,
                range=f"{unmatched_tab}!A:N",
                valueInputOption="RAW",
                insertDataOption="INSERT_ROWS",
                body={"values": unmatched_append}
            ).execute()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error appending QONTO_UNMATCHED: {e}")

    # ── STEP 8: Recalculate ALL réel from MATCH_LOG ──
    try:
        recalc_reel_from_match_log(svc, sess)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error recalculating RÉEL: {e}")

    return {
        "status":       "ok",
        "transactions": len(raw_rows),
        "new":          len(new_raw_rows),
        "duplicates":   duplicate_count,
        "matched":      results["matched"],
        "unmatched":    results["unmatched"],
        "revenue":      results["revenue"],
        "skipped":      results["skipped"],
        "summary": (
            f"✅ {len(new_raw_rows)} nouvelles transactions "
            f"({duplicate_count} déjà importées) · "
            f"{results['matched']} associées · "
            f"{results['unmatched']} non associées · "
            f"{results['revenue']} recettes · "
            f"{results['skipped']} ignorées"
        )
    }
